from pathlib import Path
from datetime import datetime
import shutil
import joblib
import warnings
import pandas as pd
import numpy as np

from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder
from sklearn.impute import SimpleImputer
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import LogisticRegression

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# =========================================================
# BACKEND PATH CONFIG
# =========================================================
try:
    from app.core.config import VMLISTE_FILE, BDD_FILE as CONFIG_BDD_FILE, OUTPUT_DIR
except Exception:
    VMLISTE_FILE = Path("vmliste_remplie.xlsx")
    CONFIG_BDD_FILE = Path("bdd_flux_maf.xlsx")
    OUTPUT_DIR = Path(".")


try:
    from app.services.job_store import find_previous_decision_by_signature
except Exception:
    find_previous_decision_by_signature = None

try:
    from app.services.job_store import find_previous_decision_loose
except Exception:
    find_previous_decision_loose = None

API_MODE = False
PENDING_DECISIONS = []
HISTORICAL_DECISIONS = []


# =========================================================
# CONFIG
# =========================================================
EXCEL_FILE = VMLISTE_FILE
BDD_FILE = CONFIG_BDD_FILE
SHEET_NAME = 0
BDD_SHEET_NAME = 0
OUTPUT_ROOT = OUTPUT_DIR / "output_basicat"
RUN_LOG_ROOT = OUTPUT_DIR / "run_logs"
BENCHMARK_CACHE_FILE = OUTPUT_DIR / "benchmark_cache.joblib"

COL_BASICAT = "BASICAT"
COL_PRODUCTION = "PRODUCTION"
COL_UTILISATION = "UTILISATION"
COL_SGIC = "SGIC"
COL_APPLICATION = "NAME"
COL_IP = "IP"
COL_APP_ID = "IDCARTO"

BDD_COLUMNS = [
    "protocol",
    "port",
    "src_ip",
    "dst_ip",
    "flowMainSG",
    "flowGrefSG",
    "direction",
    "flux",
    "Nom",
]

BDD_SIGNATURE_COLUMNS = [
    "protocol",
    "port",
    "src_ip",
    "dst_ip",
    "flowMainSG",
    "flowGrefSG",
    "direction",
]

FINAL_COLUMNS = [
    "Name",
    "Configured Service",
    "Status",
    "IP Protocol",
    "Direction",
    "port",
    "Action",
    "Configured Destination",
    "Configured Source",
    "flux",
    "Nom",
    "flowMainSG",
    "flowGrefSG",
    "applicationIdFlowMainSG",
    "applicationIdFlowGrefSG",
    "niveau_securite",
    "ml_modele",
    "ml_confiance",
    "justification",
    "anomalie",
]

# Colonnes à exclure des exports FR / SNIF destinés aux équipes externes
EXPORT_EXCLUDE_LABELS = [
    "niveau_securite",
    "ml_modele",
    "ml_confiance",
    "justification",
    "anomalie",
]

DEFAULT_STATUS = "Enabled"
DEFAULT_ACTION = "ALLOW"
DEFAULT_PROTOCOL = "TCP"
DEFAULT_DIRECTION = "outbound"

WHITE_SCORE_THRESHOLD = 100
GREY_SCORE_THRESHOLD = 60
ML_AUTO_THRESHOLD = 0.90
ML_REVIEW_THRESHOLD = 0.65
ML_FR_THRESHOLD = 0.80

FILL_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFF")
FILL_GREY = PatternFill(fill_type="solid", fgColor="BFBFBF")
FILL_BLACK = PatternFill(fill_type="solid", fgColor="000000")

FONT_BLACK = Font(color="000000")
FONT_WHITE = Font(color="FFFFFF")


# =========================================================
# INTRO
# =========================================================
def print_benchmark_intro():
    print("=" * 96)
    print("MODELE ML : exécution simplifiée sans phase de benchmark")
    print("Le moteur s'appuie directement sur la BDD et les validations humaines.")
    print("=" * 96)
    print()


def _bdd_signature(bdd_path: Path):
    if not bdd_path.exists():
        return None

    stat = bdd_path.stat()
    return {
        "path": str(bdd_path.resolve()),
        "mtime_ns": stat.st_mtime_ns,
        "size": stat.st_size,
    }


def _load_benchmark_cache(signature):
    if not BENCHMARK_CACHE_FILE.exists():
        return None

    try:
        cached = joblib.load(BENCHMARK_CACHE_FILE)
    except Exception:
        return None

    if cached.get("signature") != signature:
        return None

    best_model_info = cached.get("best_model_info")
    benchmark_df = cached.get("benchmark_df")

    if benchmark_df is None:
        benchmark_df = pd.DataFrame()

    return best_model_info, benchmark_df


def _save_benchmark_cache(signature, best_model_info, benchmark_df):
    try:
        BENCHMARK_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        joblib.dump(
            {
                "signature": signature,
                "best_model_info": best_model_info,
                "benchmark_df": benchmark_df,
            },
            BENCHMARK_CACHE_FILE,
        )
    except Exception as exc:
        print(f"Impossible d'enregistrer le cache benchmark: {exc}")


# =========================================================
# OUTILS GENERAUX
# =========================================================
def clean_value(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize(value):
    return clean_value(value).upper()


def safe_filename(name: str) -> str:
    invalid = '<>:"/\\|?*'
    out = "".join("_" if c in invalid else c for c in name)
    return out.strip() or "output"


def ask_non_empty(prompt):
    while True:
        value = input(prompt).strip()
        if value:
            return value
        print("Valeur obligatoire.")


def ask_choice(prompt, choices):
    valid = {c.lower(): c for c in choices}
    while True:
        value = input(prompt).strip().lower()
        if value in valid:
            return valid[value]
        print(f"Choix autorisés : {', '.join(choices)}")


def ask_optional(prompt):
    return input(prompt).strip()


def get_environment(production_value, utilisation_value=None):
    """
    Détermine l'environnement (prod ou horsprod) en consolidant :
    1. UTILISATION : Qualification, integration, devellopement, Métrologie, formation → horsprod
                     production, preproduction, datarecovery → prod
    2. PRODUCTION : X → prod, vide → horsprod

    Si les deux colonnes existent et divergent, UTILISATION sert de signal principal
    et PRODUCTION reste un contrôle croisé.
    """
    util_result = None
    prod_result = None

    if utilisation_value is not None:
        util_norm = normalize(utilisation_value)
        if any(keyword in util_norm for keyword in ["qualification", "integration", "devellopement", "metrologie", "formation"]):
            util_result = "horsprod"
        elif any(keyword in util_norm for keyword in ["production", "preproduction", "datarecovery"]):
            util_result = "prod"

    value = normalize(production_value)
    if value == "X":
        prod_result = "prod"
    elif clean_value(production_value) == "":
        prod_result = "horsprod"

    if util_result is not None and prod_result is not None:
        if util_result != prod_result:
            print(
                f"⚠️  Incohérence ENV détectée: UTILISATION={utilisation_value!r} -> {util_result}, "
                f"PRODUCTION={production_value!r} -> {prod_result}. UTILISATION est prioritaire."
            )
        return util_result

    if util_result is not None:
        return util_result

    if prod_result is not None:
        return prod_result

    return None


def find_column_case_insensitive(columns, target_names):
    if isinstance(target_names, str):
        target_names = [target_names]

    normalized_map = {normalize(c): c for c in columns}
    for target in target_names:
        if normalize(target) in normalized_map:
            return normalized_map[normalize(target)]
    return None


def now_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def ensure_folder(path: Path):
    path.mkdir(parents=True, exist_ok=True)
    return path


# =========================================================
# EXCEL
# =========================================================
def autosize_excel_columns(file_path: Path):
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    value = "" if cell.value is None else str(cell.value)
                    length = max(length, len(value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(length + 2, 12), 45)
    wb.save(file_path)


def apply_security_colors(file_path: Path, sheet_name=None, column_name="niveau_securite"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    headers = [clean_value(c.value) for c in ws[1]]
    if column_name not in headers:
        wb.save(file_path)
        return

    col_idx = headers.index(column_name) + 1

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = clean_value(cell.value)

        if val == "White":
            cell.fill = FILL_WHITE
            cell.font = FONT_BLACK
        elif val == "Grey":
            cell.fill = FILL_GREY
            cell.font = FONT_BLACK
        elif val == "Black":
            cell.fill = FILL_BLACK
            cell.font = FONT_WHITE

    wb.save(file_path)


def write_run_log(details_df, summary_dict, benchmark_df=None, conflicts_df=None):
    ensure_folder(RUN_LOG_ROOT)
    ts = now_timestamp()
    out_file = RUN_LOG_ROOT / f"run_log_{ts}.xlsx"

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        if details_df is not None and not details_df.empty:
            details_df.to_excel(writer, index=False, sheet_name="details")
        pd.DataFrame([summary_dict]).to_excel(writer, index=False, sheet_name="summary")

        if benchmark_df is not None and not benchmark_df.empty:
            benchmark_df.to_excel(writer, index=False, sheet_name="benchmark")

        if conflicts_df is not None and not conflicts_df.empty:
            conflicts_df.to_excel(writer, index=False, sheet_name="bdd_conflicts")

    autosize_excel_columns(out_file)
    return out_file


# =========================================================
# BDD
# =========================================================
def ensure_bdd_columns(df_bdd):
    for col in BDD_COLUMNS:
        if col not in df_bdd.columns:
            df_bdd[col] = ""
    return df_bdd[BDD_COLUMNS].copy()


def load_bdd(bdd_path: Path):
    """
    Charge la BDD avec détection automatique.
    Si le chemin ne contient rien, essaie de détecter automatiquement.
    """
    try:
        from app.services.file_detector import auto_detect_and_load_bdd
        # Essayer détection automatique d'abord
        try:
            df_bdd = auto_detect_and_load_bdd(Path("data"))
            return df_bdd
        except Exception:
            # Si détection échoue, utiliser le chemin par défaut
            pass
    except ImportError:
        pass
    
    if bdd_path.exists():
        df_bdd = pd.read_excel(bdd_path, sheet_name=BDD_SHEET_NAME, dtype=str)
        df_bdd.columns = [str(c).strip() for c in df_bdd.columns]
        for c in df_bdd.columns:
            df_bdd[c] = df_bdd[c].apply(clean_value)
        df_bdd = ensure_bdd_columns(df_bdd)
    else:
        df_bdd = pd.DataFrame(columns=BDD_COLUMNS)
    return df_bdd


def save_bdd(df_bdd, bdd_path: Path):
    df_bdd = ensure_bdd_columns(df_bdd)
    df_bdd = df_bdd.drop_duplicates().reset_index(drop=True)

    if bdd_path.exists():
        backup_dir = ensure_folder(bdd_path.parent / "bdd_backups")
        backup_file = backup_dir / f"{bdd_path.stem}_backup_{now_timestamp()}{bdd_path.suffix}"
        shutil.copy2(bdd_path, backup_file)

    df_bdd.to_excel(bdd_path, index=False)
    autosize_excel_columns(bdd_path)


def add_bdd_learning_row(
    df_bdd,
    protocol,
    port,
    src_ip,
    dst_ip,
    flow_main_sg,
    flow_gref_sg,
    direction,
    flux,
    nom,
):
    new_row = {
        "protocol": clean_value(protocol),
        "port": clean_value(port),
        "src_ip": clean_value(src_ip),
        "dst_ip": clean_value(dst_ip),
        "flowMainSG": clean_value(flow_main_sg),
        "flowGrefSG": clean_value(flow_gref_sg),
        "direction": clean_value(direction),
        "flux": clean_value(flux),
        "Nom": clean_value(nom),
    }
    return pd.concat([df_bdd, pd.DataFrame([new_row])], ignore_index=True)


def quality_check_bdd(df_bdd):
    if df_bdd.empty:
        return df_bdd, pd.DataFrame()

    df = df_bdd.copy()
    for col in BDD_COLUMNS:
        df[col] = df[col].apply(clean_value)

    df = df.drop_duplicates().reset_index(drop=True)

    grouped = (
        df.groupby(BDD_SIGNATURE_COLUMNS)[["flux", "Nom"]]
        .agg(lambda x: list(sorted(set(v for v in x if clean_value(v)))))
        .reset_index()
    )

    conflict_rows = []
    for _, row in grouped.iterrows():
        flux_values = row["flux"]
        nom_values = row["Nom"]
        if len(flux_values) > 1 or len(nom_values) > 1:
            conflict_rows.append(row)

    conflicts_df = pd.DataFrame(conflict_rows)
    return df, conflicts_df


# =========================================================
# IDCARTO
# =========================================================
def load_idcarto_mapping(excel_path: Path):
    if not excel_path.exists():
        return {}

    try:
        df_map = pd.read_excel(excel_path, sheet_name=SHEET_NAME, dtype=str)
    except Exception:
        return {}

    df_map.columns = [str(c).strip() for c in df_map.columns]

    required = [COL_IP, COL_APP_ID]
    for col in required:
        if col not in df_map.columns:
            return {}

    df_map[COL_IP] = df_map[COL_IP].apply(clean_value)
    df_map[COL_APP_ID] = df_map[COL_APP_ID].apply(clean_value)

    mapping = {}
    for _, row in df_map.iterrows():
        ip = clean_value(row.get(COL_IP, ""))
        app_id = clean_value(row.get(COL_APP_ID, ""))
        if ip and app_id and ip not in mapping:
            mapping[ip] = app_id

    return mapping


# =========================================================
# SCORE / MATCH
# =========================================================
def score_bdd_match(bdd_row, src_ip, dst_ip, protocol, port, flow_main_sg, flow_gref_sg, direction):
    score = 0

    if normalize(bdd_row.get("flowMainSG", "")) == normalize(flow_main_sg):
        score += 25
    if normalize(bdd_row.get("flowGrefSG", "")) == normalize(flow_gref_sg):
        score += 25
    if normalize(bdd_row.get("src_ip", "")) == normalize(src_ip):
        score += 15
    if normalize(bdd_row.get("dst_ip", "")) == normalize(dst_ip):
        score += 15
    if normalize(bdd_row.get("port", "")) == normalize(port):
        score += 10
    if normalize(bdd_row.get("protocol", "")) == normalize(protocol):
        score += 10

    return score


def get_confidence_color(score):
    if score >= WHITE_SCORE_THRESHOLD:
        return "White"
    elif score >= GREY_SCORE_THRESHOLD:
        return "Grey"
    return "Black"


def build_signature(protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, direction):
    return (
        clean_value(protocol),
        clean_value(port),
        clean_value(src_ip),
        clean_value(dst_ip),
        clean_value(flow_main_sg),
        clean_value(flow_gref_sg),
        clean_value(direction),
    )


def find_exact_match(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, direction):
    if df_bdd.empty:
        return None

    target_sig = build_signature(protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, direction)

    for _, row in df_bdd.iterrows():
        row_sig = build_signature(
            row.get("protocol", ""),
            row.get("port", ""),
            row.get("src_ip", ""),
            row.get("dst_ip", ""),
            row.get("flowMainSG", ""),
            row.get("flowGrefSG", ""),
            row.get("direction", ""),
        )
        if row_sig == target_sig:
            return row
    return None


def find_best_match(df_bdd, src_ip, dst_ip, protocol, port, flow_main_sg, flow_gref_sg, direction):
    if df_bdd.empty:
        return None, 0

    scored_rows = []
    for idx, row in df_bdd.iterrows():
        score = score_bdd_match(
            bdd_row=row,
            src_ip=src_ip,
            dst_ip=dst_ip,
            protocol=protocol,
            port=port,
            flow_main_sg=flow_main_sg,
            flow_gref_sg=flow_gref_sg,
            direction=direction,
        )
        scored_rows.append((idx, score))

    scored_rows.sort(key=lambda x: x[1], reverse=True)
    best_idx, best_score = scored_rows[0]

    if best_score <= 0:
        return None, 0

    return df_bdd.loc[best_idx], best_score


def infer_direction(row):
    existing = clean_value(row.get("Direction", "")) or clean_value(row.get("direction", ""))
    if existing:
        val = existing.lower()
        if val in {"inbound", "outbound"}:
            return val

    src_sg = clean_value(row.get("flowMainSG", ""))
    dst_sg = clean_value(row.get("flowGrefSG", ""))

    if src_sg and dst_sg:
        return "outbound"

    return DEFAULT_DIRECTION


def infer_name(application, flow_main_sg, flow_gref_sg, direction, flux):
    parts = [
        clean_value(application),
        clean_value(flux),
        clean_value(direction),
        clean_value(flow_main_sg),
        clean_value(flow_gref_sg),
    ]
    parts = [p for p in parts if p]
    return "_".join(parts) if parts else "FLOW_RULE"


# =========================================================
# BENCHMARK ML
# =========================================================
def get_model_dict():
    return {
        "RandomForest": RandomForestClassifier(
            n_estimators=250,
            random_state=42,
            class_weight="balanced",
        ),
        "GradientBoosting": GradientBoostingClassifier(random_state=42),
        "DecisionTree": DecisionTreeClassifier(
            random_state=42,
            class_weight="balanced",
        ),
        "KNN": KNeighborsClassifier(n_neighbors=3),
        "LogisticRegression": LogisticRegression(
            max_iter=2000,
            class_weight="balanced",
        ),
    }


def build_preprocessor():
    categorical_features = ["protocol", "port", "src_ip", "dst_ip", "flowMainSG", "flowGrefSG"]

    return ColumnTransformer(
        transformers=[
            (
                "cat",
                Pipeline(
                    steps=[
                        ("imputer", SimpleImputer(strategy="constant", fill_value="missing")),
                        ("onehot", OneHotEncoder(handle_unknown="ignore")),
                    ]
                ),
                categorical_features,
            )
        ]
    )


def run_ml_benchmark(df_bdd):
    bdd_signature = _bdd_signature(Path(BDD_FILE))
    cached_result = _load_benchmark_cache(bdd_signature)

    if cached_result is not None:
        print("BENCHMARK ML - cache utilise, aucun recalcul.")
        print()
        return cached_result

    print("=" * 96)
    print("BENCHMARK ML - SELECTION DU MODELE")
    print("=" * 96)

    if df_bdd.empty or "flux" not in df_bdd.columns:
        print("BDD vide ou colonne 'flux' absente. Benchmark impossible.")
        print("Fallback : moteur hybride BDD + validation humaine.")
        print("=" * 96)
        print()
        return None, pd.DataFrame()

    df_ml = df_bdd.copy()
    required_cols = ["protocol", "port", "src_ip", "dst_ip", "flowMainSG", "flowGrefSG", "flux"]
    for col in required_cols:
        if col not in df_ml.columns:
            print(f"Colonne manquante pour benchmark : {col}")
            print("Fallback : moteur hybride BDD + validation humaine.")
            print("=" * 96)
            print()
            return None, pd.DataFrame()

    for col in required_cols:
        df_ml[col] = df_ml[col].apply(clean_value)

    df_ml = df_ml[df_ml["flux"] != ""].copy()

    if len(df_ml) < 15:
        print(f"Pas assez de données pour benchmark fiable ({len(df_ml)} lignes).")
        print("Fallback : moteur hybride BDD + validation humaine.")
        print("=" * 96)
        print()
        return None, pd.DataFrame()

    if df_ml["flux"].nunique() < 2:
        print("Une seule classe 'flux' présente dans la BDD.")
        print("Fallback : moteur hybride BDD + validation humaine.")
        print("=" * 96)
        print()
        return None, pd.DataFrame()

    X = df_ml[["protocol", "port", "src_ip", "dst_ip", "flowMainSG", "flowGrefSG"]]
    y = df_ml["flux"]

    n_splits = min(5, max(2, y.value_counts().min()))
    cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)

    preprocessor = build_preprocessor()
    models = get_model_dict()
    warnings.filterwarnings("ignore")

    results = []

    for model_name, model in models.items():
        try:
            pipeline = Pipeline(
                steps=[
                    ("preprocessor", preprocessor),
                    ("model", model),
                ]
            )

            scores = cross_validate(
                pipeline,
                X,
                y,
                cv=cv,
                scoring={
                    "accuracy": "accuracy",
                    "precision": "precision_weighted",
                    "recall": "recall_weighted",
                    "f1": "f1_weighted",
                },
                error_score="raise",
            )

            result = {
                "model": model_name,
                "accuracy": round(np.mean(scores["test_accuracy"]), 4),
                "precision": round(np.mean(scores["test_precision"]), 4),
                "recall": round(np.mean(scores["test_recall"]), 4),
                "f1": round(np.mean(scores["test_f1"]), 4),
            }
            results.append(result)

        except Exception as e:
            print(f"Erreur benchmark pour {model_name}: {e}")

    if not results:
        print("Aucun modèle benchmarkable.")
        print("Fallback : moteur hybride BDD + validation humaine.")
        print("=" * 96)
        print()
        return None, pd.DataFrame()

    benchmark_df = pd.DataFrame(results)
    benchmark_df = benchmark_df.sort_values(
        by=["f1", "recall", "accuracy"],
        ascending=False
    ).reset_index(drop=True)

    print("Résultats benchmark :")
    for _, r in benchmark_df.iterrows():
        print(
            f"- {r['model']}: "
            f"accuracy={r['accuracy']:.4f} | "
            f"precision={r['precision']:.4f} | "
            f"recall={r['recall']:.4f} | "
            f"f1={r['f1']:.4f}"
        )

    best_model_name = benchmark_df.iloc[0]["model"]
    print("\nModèle retenu :")
    print(f"- {best_model_name}")
    print(f"- accuracy={benchmark_df.iloc[0]['accuracy']:.4f}")
    print(f"- precision={benchmark_df.iloc[0]['precision']:.4f}")
    print(f"- recall={benchmark_df.iloc[0]['recall']:.4f}")
    print(f"- f1-score={benchmark_df.iloc[0]['f1']:.4f}")
    print()
    print("Critère de choix : F1-score, puis recall, puis accuracy.")
    print("=" * 96)
    print()

    best_model = get_model_dict()[best_model_name]
    final_pipeline = Pipeline(
        steps=[
            ("preprocessor", preprocessor),
            ("model", best_model),
        ]
    )
    final_pipeline.fit(X, y)

    best = {
        "model": best_model_name,
        "pipeline": final_pipeline,
        "accuracy": float(benchmark_df.iloc[0]["accuracy"]),
        "precision": float(benchmark_df.iloc[0]["precision"]),
        "recall": float(benchmark_df.iloc[0]["recall"]),
        "f1": float(benchmark_df.iloc[0]["f1"]),
    }

    if bdd_signature is not None:
        _save_benchmark_cache(bdd_signature, best, benchmark_df)

    return best, benchmark_df


def predict_flux_with_ml(best_model_info, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg):
    if best_model_info is None:
        return None, 0.0

    pipeline = best_model_info.get("pipeline")

    if pipeline is None:
        return None, 0.0

    port_num = pd.to_numeric(clean_value(port), errors="coerce")

    X_new = pd.DataFrame([{
        "port": port_num,
        "protocol": clean_value(protocol) or DEFAULT_PROTOCOL,
        "src_ip": clean_value(src_ip),
        "dst_ip": clean_value(dst_ip),
        "flowMainSG": clean_value(flow_main_sg),
        "flowGrefSG": clean_value(flow_gref_sg),
    }])

    try:
        pred = pipeline.predict(X_new)[0]
        confidence = 0.0

        if hasattr(pipeline, "predict_proba"):
            probs = pipeline.predict_proba(X_new)[0]
            confidence = float(np.max(probs))

        return clean_value(pred), confidence

    except Exception as exc:
        print(f"[ML DEBUG] Erreur prediction ML: {exc}")
        return None, 0.0


def find_nom_from_bdd(df_bdd, predicted_flux, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg):
    if df_bdd.empty or not predicted_flux:
        return ""

    candidates = df_bdd[df_bdd["flux"].apply(normalize) == normalize(predicted_flux)].copy()
    if candidates.empty:
        return ""

    best_match, _ = find_best_match(
        df_bdd=candidates,
        src_ip=src_ip,
        dst_ip=dst_ip,
        protocol=protocol,
        port=port,
        flow_main_sg=flow_main_sg,
        flow_gref_sg=flow_gref_sg,
        direction=""
    )

    if best_match is None:
        return ""

    return clean_value(best_match.get("Nom", ""))


def compute_suggestion(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg):
    """
    Retourne une suggestion (flux, nom) basée sur la BDD en utilisant
    un meilleur match ou des candidats pour la signature fournie.
    """
    if df_bdd is None or df_bdd.empty:
        return "", ""

    # 1) exact match
    exact = find_exact_match(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, "")
    if exact is not None:
        return clean_value(exact.get("flux", "")), clean_value(exact.get("Nom", ""))

    # 2) best match
    best_match, best_score = find_best_match(df_bdd, src_ip, dst_ip, protocol, port, flow_main_sg, flow_gref_sg, "")
    if best_match is not None and best_score > 0:
        return clean_value(best_match.get("flux", "")), clean_value(best_match.get("Nom", ""))

    # 3) candidates majority
    candidates = candidate_rows_for_pattern(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg)
    if not candidates.empty:
        # pick most frequent flux/nom
        flux_vals = [v for v in candidates["flux"].apply(clean_value) if v]
        nom_vals = [v for v in candidates["Nom"].apply(clean_value) if v]
        if flux_vals:
            from collections import Counter

            flux = Counter(flux_vals).most_common(1)[0][0]
        else:
            flux = ""
        if nom_vals:
            nom = Counter(nom_vals).most_common(1)[0][0]
        else:
            nom = ""
        return clean_value(flux), clean_value(nom)

    return "", ""


# =========================================================
# MOTEUR COMMUN FR / SNIF
# =========================================================
def is_confident_bdd_pattern(df_bdd, src_ip, flow_main_sg):
    """
    Auto-validation BDD par pattern stable.
    Si la BDD contient toujours le même couple flux/Nom pour le même
    src_ip + flowMainSG, on considère que le pattern est connu.
    """
    if df_bdd.empty:
        return False, None

    src_ip_n = normalize(src_ip)
    flow_main_sg_n = normalize(flow_main_sg)

    df_filtered = df_bdd[
        (df_bdd["src_ip"].apply(normalize) == src_ip_n) &
        (df_bdd["flowMainSG"].apply(normalize) == flow_main_sg_n)
    ].copy()

    if df_filtered.empty:
        return False, None

    unique_flux = [v for v in df_filtered["flux"].apply(clean_value).unique() if v]
    unique_nom = [v for v in df_filtered["Nom"].apply(clean_value).unique() if v]

    if len(unique_flux) == 1 and len(unique_nom) == 1:
        return True, df_filtered.iloc[0]

    return False, None



def get_dynamic_bdd_threshold(df_bdd):
    n = 0 if df_bdd is None else len(df_bdd)
    if n < 30:
        return 75
    if n < 100:
        return 65
    return 55


def get_dynamic_ml_threshold(df_bdd, benchmark_info=None):
    f1 = 0.0
    if benchmark_info:
        try:
            f1 = float(benchmark_info.get("f1", 0.0))
        except Exception:
            f1 = 0.0
    n = 0 if df_bdd is None else len(df_bdd)
    if f1 >= 0.90 and n >= 100:
        return 0.88
    if f1 >= 0.80 and n >= 50:
        return 0.92
    return 0.95


def get_model_name(best_model_info):
    """Retourne le nom du modèle ML sans faire planter le moteur si l’objet est None ou incomplet."""
    if not best_model_info:
        return ""

    if isinstance(best_model_info, dict):
        return clean_value(best_model_info.get("model") or best_model_info.get("model_name") or "")

    return clean_value(getattr(best_model_info, "model", "") or getattr(best_model_info, "model_name", ""))


def detect_anomalies(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, direction):
    anomalies = []
    if df_bdd is None or df_bdd.empty:
        return anomalies
    same_sig = df_bdd[(df_bdd["protocol"].apply(normalize) == normalize(protocol)) & (df_bdd["port"].apply(normalize) == normalize(port)) & (df_bdd["src_ip"].apply(normalize) == normalize(src_ip)) & (df_bdd["dst_ip"].apply(normalize) == normalize(dst_ip)) & (df_bdd["flowMainSG"].apply(normalize) == normalize(flow_main_sg)) & (df_bdd["flowGrefSG"].apply(normalize) == normalize(flow_gref_sg))]
    if not same_sig.empty:
        if same_sig["flux"].apply(clean_value).nunique() > 1:
            anomalies.append("Conflit BDD: même signature avec plusieurs flux")
        if same_sig["Nom"].apply(clean_value).nunique() > 1:
            anomalies.append("Conflit BDD: même signature avec plusieurs noms")
    same_main = df_bdd[(df_bdd["src_ip"].apply(normalize) == normalize(src_ip)) & (df_bdd["flowMainSG"].apply(normalize) == normalize(flow_main_sg))]
    if not same_main.empty:
        if same_main["flux"].apply(clean_value).nunique() > 1:
            anomalies.append("Pattern instable: src_ip + flowMainSG associés à plusieurs flux")
        if same_main["Nom"].apply(clean_value).nunique() > 1:
            anomalies.append("Pattern instable: src_ip + flowMainSG associés à plusieurs noms")
    same_service = df_bdd[(df_bdd["protocol"].apply(normalize) == normalize(protocol)) & (df_bdd["port"].apply(normalize) == normalize(port))]
    if len(same_service) >= 5 and same_service["flux"].apply(clean_value).nunique() > 2:
        anomalies.append("Service ambigu: port/protocol associés à plusieurs types de flux")
    return anomalies


def candidate_rows_for_pattern(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg):
    if df_bdd is None or df_bdd.empty:
        return pd.DataFrame(columns=BDD_COLUMNS)
    mask = pd.Series([True] * len(df_bdd), index=df_bdd.index)
    for col, value in [("protocol", protocol), ("port", port), ("src_ip", src_ip), ("dst_ip", dst_ip), ("flowMainSG", flow_main_sg), ("flowGrefSG", flow_gref_sg)]:
        if clean_value(value):
            mask &= df_bdd[col].apply(normalize) == normalize(value)
    return df_bdd[mask].copy()


def has_unique_candidate_result(df_candidates):
    if df_candidates is None or df_candidates.empty:
        return False
    flux_values = [v for v in df_candidates["flux"].apply(clean_value).unique() if v]
    nom_values = [v for v in df_candidates["Nom"].apply(clean_value).unique() if v]
    return len(flux_values) == 1 and len(nom_values) == 1


def build_justification(decision_source, flux, nom, score=None, ml_conf=None, anomalies=None):
    parts = [f"source={decision_source}", f"flux={clean_value(flux)}", f"Nom={clean_value(nom)}"]
    if score not in (None, ""):
        parts.append(f"score_bdd={score}")
    if ml_conf not in (None, ""):
        parts.append(f"ml_conf={ml_conf}")
    if anomalies:
        parts.append("anomalies=" + " | ".join(anomalies))
    return "; ".join(parts)


def pack_result(flux, nom, niveau_securite="White", ml_modele="", ml_confiance="", dst_ip="", flow_gref_sg="", port="", direction="", learned=False, decision_source="", justification="", anomalie=""):
    result = {
        "flux": clean_value(flux),
        "nom": clean_value(nom),
        "niveau_securite": clean_value(niveau_securite),
        "ml_modele": clean_value(ml_modele),
        "ml_confiance": clean_value(ml_confiance),
        "dst_ip": clean_value(dst_ip),
        "dstIp": clean_value(dst_ip),
        "destination_ip": clean_value(dst_ip),
        "flow_gref_sg": clean_value(flow_gref_sg),
        "flowGrefSG": clean_value(flow_gref_sg),
        "sg_cible": clean_value(flow_gref_sg),
        "port": clean_value(port),
        "direction": clean_value(direction) or DEFAULT_DIRECTION,
        "learned": bool(learned),
        "decision_source": clean_value(decision_source),
        "justification": clean_value(justification),
        "anomalie": clean_value(anomalie),
    }
    return result


def add_pending_decision(
    env_name,
    source,
    score,
    protocol,
    port,
    src_ip,
    dst_ip,
    flow_main_sg,
    flow_gref_sg,
    direction,
    proposed_flux,
    proposed_nom,
    suggested_flux="",
    suggested_nom="",
    ml_modele="",
    ml_confiance="",
    seuil_auto="",
    anomalie="",
):
    decision_id = f"{env_name}-{len(PENDING_DECISIONS) + 1:05d}"
    payload = {
        "decision_id": decision_id,
        "env": clean_value(env_name),
        "source": clean_value(source),
        "score": clean_value(score),
        "protocol": clean_value(protocol),
        "port": clean_value(port),
        "src_ip": clean_value(src_ip),
        "dst_ip": clean_value(dst_ip),
        "dstIp": clean_value(dst_ip),
        "destination_ip": clean_value(dst_ip),
        "flowMainSG": clean_value(flow_main_sg),
        "flowGrefSG": clean_value(flow_gref_sg),
        "flow_gref_sg": clean_value(flow_gref_sg),
        "sg_cible": clean_value(flow_gref_sg),
        "direction": clean_value(direction),
        "proposed_flux": clean_value(proposed_flux),
        "proposed_nom": clean_value(proposed_nom),
        "suggested_flux": clean_value(suggested_flux),
        "suggested_nom": clean_value(suggested_nom),
        "ml_modele": clean_value(ml_modele),
        "ml_confiance": clean_value(ml_confiance),
        "seuil_auto": clean_value(seuil_auto),
        "anomalie": clean_value(anomalie),
    }
    PENDING_DECISIONS.append(payload)
    return decision_id


def add_historical_decision(
    env_name,
    source,
    protocol,
    port,
    src_ip,
    dst_ip,
    flow_main_sg,
    flow_gref_sg,
    direction,
    proposed_flux,
    proposed_nom,
    previous_score="",
    previous_ml_modele="",
    previous_ml_confiance="",
    previous_seuil_auto="",
    anomalie="",
):
    """
    Ajoute une ligne déjà connue dans un bloc séparé.

    Contrairement à PENDING_DECISIONS, ces lignes ne bloquent pas le workflow :
    elles sont déjà considérées comme validées via l'historique.
    L'utilisateur peut les corriger optionnellement côté interface.
    """
    decision_id = f"{env_name}-hist-{len(HISTORICAL_DECISIONS) + 1:05d}"

    payload = {
        "decision_id": decision_id,
        "env": clean_value(env_name),
        "source": clean_value(source) or "HISTORICAL_AUTO",
        "score": "100",
        "historical_score": clean_value(previous_score),
        "protocol": clean_value(protocol),
        "port": clean_value(port),
        "src_ip": clean_value(src_ip),
        "dst_ip": clean_value(dst_ip),
        "dstIp": clean_value(dst_ip),
        "destination_ip": clean_value(dst_ip),
        "flowMainSG": clean_value(flow_main_sg),
        "flowGrefSG": clean_value(flow_gref_sg),
        "flow_gref_sg": clean_value(flow_gref_sg),
        "sg_cible": clean_value(flow_gref_sg),
        "direction": clean_value(direction) or DEFAULT_DIRECTION,
        "proposed_flux": clean_value(proposed_flux),
        "proposed_nom": clean_value(proposed_nom),
        "suggested_flux": clean_value(proposed_flux),
        "suggested_nom": clean_value(proposed_nom),
        "final_flux": clean_value(proposed_flux),
        "final_nom": clean_value(proposed_nom),
        "ml_modele": clean_value(previous_ml_modele),
        # Historique = déjà validé auparavant, donc confiance affichée à 100%.
        # L'ancienne confiance ML reste disponible dans previous_ml_confiance.
        "ml_confiance": "1.0",
        "previous_ml_confiance": clean_value(previous_ml_confiance),
        "seuil_auto": clean_value(previous_seuil_auto),
        "anomalie": clean_value(anomalie),
        "already_validated": True,
        "requires_validation": False,
        "can_correct": True,
    }

    HISTORICAL_DECISIONS.append(payload)
    return decision_id

def resolve_flow_with_bdd_and_ml(
    df_bdd,
    protocol,
    port,
    src_ip,
    dst_ip,
    flow_main_sg,
    flow_gref_sg,
    direction,
    application,
    env_name,
    basicat_code="",
    best_model_info=None,
    ask_user=True,
):
    bdd_threshold = get_dynamic_bdd_threshold(df_bdd)
    ml_threshold = get_dynamic_ml_threshold(df_bdd, best_model_info)
    anomalies = detect_anomalies(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, direction)
    anomalie_txt = " | ".join(anomalies)

    # =====================================================
    # HISTORIQUE BASICAT : ligne déjà validée auparavant
    # 1) recherche stricte par signature complète
    # 2) fallback souple pour FR, car port/dst/SG cible peuvent être vides
    # =====================================================
    previous_decision = None

    if clean_value(basicat_code) and clean_value(env_name):
        if find_previous_decision_by_signature is not None:
            try:
                previous_decision = find_previous_decision_by_signature(
                    basicat=basicat_code,
                    env=env_name,
                    protocol=protocol,
                    port=port,
                    src_ip=src_ip,
                    dst_ip=dst_ip,
                    flowMainSG=flow_main_sg,
                    flowGrefSG=flow_gref_sg,
                    direction=direction,
                )
            except Exception as exc:
                print(f"[HISTORICAL DEBUG] Recherche stricte impossible: {exc}")
                previous_decision = None

        if previous_decision is None and find_previous_decision_loose is not None:
            try:
                # Recherche souple FR : le FR n’a pas toujours encore port/dst_ip/SG cible.
                previous_decision = find_previous_decision_loose(
                    basicat=basicat_code,
                    env=env_name,
                    src_ip=src_ip,
                    flowMainSG=flow_main_sg,
                )
            except Exception as exc:
                print(f"[HISTORICAL DEBUG] Recherche souple impossible: {exc}")
                previous_decision = None

    if previous_decision is not None and not ask_user:
        previous_flux = clean_value(previous_decision.get("final_flux", ""))
        previous_nom = clean_value(previous_decision.get("final_nom", ""))
        previous_score = clean_value(previous_decision.get("score", ""))
        previous_ml_modele = clean_value(previous_decision.get("ml_modele", ""))
        previous_ml_confiance = clean_value(previous_decision.get("ml_confiance", ""))
        previous_seuil_auto = clean_value(previous_decision.get("seuil_auto", ""))

        # Pour FR, ces valeurs sont souvent vides dans la nouvelle ligne.
        # On réutilise donc celles de l'historique pour les afficher au front.
        historical_port = clean_value(port) or clean_value(previous_decision.get("port", ""))
        historical_dst_ip = clean_value(dst_ip) or clean_value(previous_decision.get("dst_ip", ""))
        historical_flow_gref = clean_value(flow_gref_sg) or clean_value(previous_decision.get("flowGrefSG", ""))
        historical_direction = clean_value(direction) or DEFAULT_DIRECTION

        add_historical_decision(
            env_name=env_name,
            source="HISTORICAL_AUTO",
            protocol=protocol,
            port=historical_port,
            src_ip=src_ip,
            dst_ip=historical_dst_ip,
            flow_main_sg=flow_main_sg,
            flow_gref_sg=historical_flow_gref,
            direction=historical_direction,
            proposed_flux=previous_flux,
            proposed_nom=previous_nom,
            previous_score=previous_score,
            previous_ml_modele=previous_ml_modele,
            previous_ml_confiance=previous_ml_confiance,
            previous_seuil_auto=previous_seuil_auto,
            anomalie=anomalie_txt,
        )

        return pack_result(
            previous_flux,
            previous_nom,
            niveau_securite="White",
            ml_modele=previous_ml_modele,
            # Historique déjà validé : affichage à 100%.
            ml_confiance="1.0",
            dst_ip=historical_dst_ip,
            flow_gref_sg=historical_flow_gref,
            port=historical_port,
            direction=historical_direction,
            learned=False,
            decision_source="HISTORICAL_AUTO",
            justification=build_justification(
                "HISTORICAL_AUTO",
                previous_flux,
                previous_nom,
                score=100,
                ml_conf="1.0",
                anomalies=anomalies,
            ),
            anomalie=anomalie_txt,
        )

    exact = find_exact_match(df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg, direction)
    if exact is not None and not anomalies:
        flux = clean_value(exact.get("flux", ""))
        nom = clean_value(exact.get("Nom", ""))

        resolved_dst_ip = clean_value(dst_ip) or clean_value(exact.get("dst_ip", ""))
        resolved_flow_gref = clean_value(flow_gref_sg) or clean_value(exact.get("flowGrefSG", ""))
        resolved_port = clean_value(port) or clean_value(exact.get("port", ""))
        resolved_direction = clean_value(direction) or clean_value(exact.get("direction", "")) or DEFAULT_DIRECTION

        # En API mode, les matchs exacts BDD doivent aussi être visibles dans le bloc
        # "Lignes déjà connues" afin que l'utilisateur puisse les corriger optionnellement.
        # Sans ça, SNIF passe directement à done/snif_complete et rien ne s'affiche.
        if not ask_user:
            add_historical_decision(
                env_name=env_name,
                source="BDD_EXACT_AUTO",
                protocol=protocol,
                port=resolved_port,
                src_ip=src_ip,
                dst_ip=resolved_dst_ip,
                flow_main_sg=flow_main_sg,
                flow_gref_sg=resolved_flow_gref,
                direction=resolved_direction,
                proposed_flux=flux,
                proposed_nom=nom,
                previous_score="100",
                previous_ml_modele=get_model_name(best_model_info),
                previous_ml_confiance="1.0",
                previous_seuil_auto=ml_threshold,
                anomalie=anomalie_txt,
            )

        return pack_result(
            flux, nom,
            dst_ip=resolved_dst_ip,
            flow_gref_sg=resolved_flow_gref,
            port=resolved_port,
            direction=resolved_direction,
            ml_modele=get_model_name(best_model_info),
            ml_confiance="1.0",
            decision_source="BDD_EXACT",
            justification=build_justification("BDD_EXACT", flux, nom, score=100),
        )

    confident, row_conf = is_confident_bdd_pattern(df_bdd, src_ip, flow_main_sg)
    if confident and not anomalies:
        flux = clean_value(row_conf.get("flux", ""))
        nom = clean_value(row_conf.get("Nom", ""))

        resolved_dst_ip = clean_value(dst_ip) or clean_value(row_conf.get("dst_ip", ""))
        resolved_flow_gref = clean_value(flow_gref_sg) or clean_value(row_conf.get("flowGrefSG", ""))
        resolved_port = clean_value(port) or clean_value(row_conf.get("port", ""))
        resolved_direction = clean_value(direction) or clean_value(row_conf.get("direction", "")) or DEFAULT_DIRECTION

        # Même principe pour les patterns BDD très confiants : on les affiche comme
        # lignes connues, sans validation obligatoire, avec correction optionnelle.
        if not ask_user:
            add_historical_decision(
                env_name=env_name,
                source="BDD_PATTERN_AUTO",
                protocol=protocol,
                port=resolved_port,
                src_ip=src_ip,
                dst_ip=resolved_dst_ip,
                flow_main_sg=flow_main_sg,
                flow_gref_sg=resolved_flow_gref,
                direction=resolved_direction,
                proposed_flux=flux,
                proposed_nom=nom,
                previous_score="pattern",
                previous_ml_modele=get_model_name(best_model_info),
                previous_ml_confiance="1.0",
                previous_seuil_auto=ml_threshold,
                anomalie=anomalie_txt,
            )

        return pack_result(
            flux, nom,
            dst_ip=resolved_dst_ip,
            flow_gref_sg=resolved_flow_gref,
            port=resolved_port,
            direction=resolved_direction,
            ml_modele=get_model_name(best_model_info),
            ml_confiance="1.0",
            decision_source="BDD_PATTERN_AUTO",
            justification=build_justification("BDD_PATTERN_AUTO", flux, nom, score="pattern"),
        )

    best_match, best_score = find_best_match(
        df_bdd, src_ip, dst_ip, protocol, port, flow_main_sg, flow_gref_sg, direction
    )

    candidates = candidate_rows_for_pattern(
        df_bdd, protocol, port, src_ip, dst_ip, flow_main_sg, flow_gref_sg
    )
    unique_candidate = has_unique_candidate_result(candidates)

    fallback_row = best_match
    if fallback_row is None and confident and row_conf is not None:
        fallback_row = row_conf
    if fallback_row is None and not candidates.empty:
        fallback_row = candidates.iloc[0]

    fallback_dst_ip = clean_value(dst_ip)
    fallback_flow_gref = clean_value(flow_gref_sg)
    fallback_port = clean_value(port)
    fallback_direction = clean_value(direction) or DEFAULT_DIRECTION

    if fallback_row is not None:
        if not fallback_dst_ip:
            fallback_dst_ip = clean_value(fallback_row.get("dst_ip", ""))
        if not fallback_flow_gref:
            fallback_flow_gref = clean_value(fallback_row.get("flowGrefSG", ""))
        if not fallback_port:
            fallback_port = clean_value(fallback_row.get("port", ""))
        if not fallback_direction:
            fallback_direction = clean_value(fallback_row.get("direction", "")) or DEFAULT_DIRECTION

    if best_match is not None and best_score >= bdd_threshold and unique_candidate and not anomalies:
        flux = clean_value(best_match.get("flux", ""))
        nom = clean_value(best_match.get("Nom", ""))
        return pack_result(
            flux, nom,
            dst_ip=fallback_dst_ip,
            flow_gref_sg=fallback_flow_gref,
            port=fallback_port,
            direction=fallback_direction,
            decision_source="BDD_PARTIAL_AUTO",
            justification=build_justification("BDD_PARTIAL_AUTO", flux, nom, score=best_score),
        )

    if best_match is not None and best_score >= bdd_threshold:
        proposed_flux = clean_value(best_match.get("flux", ""))
        proposed_nom = clean_value(best_match.get("Nom", ""))

        if ask_user:
            confirm = ask_choice("Valider cette proposition BDD ? (oui/non) : ", ["oui", "non"])
            if confirm == "oui":
                return pack_result(
                    proposed_flux, proposed_nom,
                    niveau_securite="Grey" if anomalies else "White",
                    dst_ip=fallback_dst_ip,
                    flow_gref_sg=fallback_flow_gref,
                    port=fallback_port,
                    direction=fallback_direction,
                    decision_source="BDD_PARTIAL_VALIDATED",
                    justification=build_justification(
                        "BDD_PARTIAL_VALIDATED",
                        proposed_flux,
                        proposed_nom,
                        score=best_score,
                        anomalies=anomalies,
                    ),
                    anomalie=anomalie_txt,
                )

            flux = ask_choice("Entrer flux (client/usine/service) : ", ["client", "usine", "service"])
            nom = ask_non_empty("Entrer Nom : ")
            final_dst_ip = fallback_dst_ip or ask_optional("Entrer Destination IP (optionnel) : ")
            final_flow_gref = fallback_flow_gref or ask_optional("Entrer flowGrefSG (optionnel) : ")

            return pack_result(
                flux, nom,
                dst_ip=final_dst_ip,
                flow_gref_sg=final_flow_gref,
                port=fallback_port,
                direction=fallback_direction,
                learned=True,
                decision_source="BDD_PARTIAL_CORRECTED",
                justification=build_justification(
                    "BDD_PARTIAL_CORRECTED",
                    flux,
                    nom,
                    score=best_score,
                    anomalies=anomalies,
                ),
                anomalie=anomalie_txt,
            )

        add_pending_decision(
            env_name,
            "BDD_PARTIAL_REVIEW",
            best_score,
            protocol,
            fallback_port,
            src_ip,
            fallback_dst_ip,
            flow_main_sg,
            fallback_flow_gref,
            fallback_direction,
            proposed_flux,
            proposed_nom,
            suggested_flux=proposed_flux,
            suggested_nom=proposed_nom,
            anomalie=anomalie_txt,
        )

        return pack_result(
            proposed_flux, proposed_nom,
            niveau_securite="Grey",
            dst_ip=fallback_dst_ip,
            flow_gref_sg=fallback_flow_gref,
            port=fallback_port,
            direction=fallback_direction,
            learned=False,
            decision_source="PENDING_BDD_REVIEW",
            justification=build_justification(
                "PENDING_BDD_REVIEW",
                proposed_flux,
                proposed_nom,
                score=best_score,
                anomalies=anomalies,
            ),
            anomalie=anomalie_txt,
        )

    if best_model_info is None:
        try:
            model_file = Path(__file__).parent.parent.parent / "models" / "rf_model.joblib"
            if model_file.exists():
                loaded = joblib.load(model_file)
                best_model_info = {"model": "rf_model", "pipeline": loaded}
        except Exception:
            best_model_info = None

    predicted_flux, predicted_conf = predict_flux_with_ml(
        best_model_info,
        protocol,
        port,
        src_ip,
        dst_ip,
        flow_main_sg,
        flow_gref_sg,
    )

    predicted_nom = find_nom_from_bdd(
        df_bdd,
        predicted_flux,
        protocol,
        port,
        src_ip,
        dst_ip,
        flow_main_sg,
        flow_gref_sg,
    )

    final_dst_ip = fallback_dst_ip
    final_flow_gref = fallback_flow_gref
    final_port = fallback_port
    final_direction = fallback_direction

    if predicted_flux and predicted_conf >= ml_threshold and not anomalies:
        nom = predicted_nom if predicted_nom else infer_name(
            application,
            flow_main_sg,
            final_flow_gref,
            final_direction,
            predicted_flux,
        )

        return pack_result(
            predicted_flux,
            nom,
            ml_modele=get_model_name(best_model_info),
            ml_confiance=str(round(predicted_conf, 4)),
            dst_ip=final_dst_ip,
            flow_gref_sg=final_flow_gref,
            port=final_port,
            direction=final_direction,
            learned=True,
            decision_source="ML_AUTO_LEARNED",
            justification=build_justification(
                "ML_AUTO_LEARNED",
                predicted_flux,
                nom,
                ml_conf=round(predicted_conf, 4),
            ),
        )

    if predicted_flux:
        if ask_user:
            confirm = ask_choice("Valider cette prédiction ? (oui/non) : ", ["oui", "non"])
            if confirm == "oui":
                nom = predicted_nom if predicted_nom else ask_non_empty("Nom (proposé vide) : ")

                return pack_result(
                    predicted_flux,
                    nom,
                    niveau_securite="Grey" if anomalies else "White",
                    ml_modele=get_model_name(best_model_info),
                    ml_confiance=str(round(predicted_conf, 4)),
                    dst_ip=final_dst_ip,
                    flow_gref_sg=final_flow_gref,
                    port=final_port,
                    direction=final_direction,
                    learned=True,
                    decision_source="ML_VALIDATED",
                    justification=build_justification(
                        "ML_VALIDATED",
                        predicted_flux,
                        nom,
                        ml_conf=round(predicted_conf, 4),
                        anomalies=anomalies,
                    ),
                    anomalie=anomalie_txt,
                )

        add_pending_decision(
            env_name,
            "ML_REVIEW",
            "",
            protocol,
            final_port,
            src_ip,
            final_dst_ip,
            flow_main_sg,
            final_flow_gref,
            final_direction,
            predicted_flux,
            predicted_nom,
            suggested_flux=predicted_flux,
            suggested_nom=predicted_nom,
            ml_modele=get_model_name(best_model_info),
            ml_confiance=str(round(predicted_conf, 4)),
            seuil_auto=str(round(ml_threshold, 4)),
            anomalie=anomalie_txt,
        )

        return pack_result(
            predicted_flux,
            predicted_nom,
            niveau_securite="Grey",
            ml_modele=get_model_name(best_model_info),
            ml_confiance=str(round(predicted_conf, 4)),
            dst_ip=final_dst_ip,
            flow_gref_sg=final_flow_gref,
            port=final_port,
            direction=final_direction,
            learned=False,
            decision_source="PENDING_ML_REVIEW",
            justification=build_justification(
                "PENDING_ML_REVIEW",
                predicted_flux,
                predicted_nom,
                ml_conf=round(predicted_conf, 4),
                anomalies=anomalies,
            ),
            anomalie=anomalie_txt,
        )

    if ask_user:
        flux = ask_choice("Entrer flux (client/usine/service) : ", ["client", "usine", "service"])
        nom = ask_non_empty("Entrer Nom : ")

        final_dst_ip = final_dst_ip or ask_optional("Entrer Destination IP (optionnel) : ")
        final_flow_gref = final_flow_gref or ask_optional("Entrer flowGrefSG (optionnel) : ")

        return pack_result(
            flux,
            nom,
            ml_modele=get_model_name(best_model_info),
            ml_confiance=str(round(predicted_conf, 4)) if predicted_conf else "",
            dst_ip=final_dst_ip,
            flow_gref_sg=final_flow_gref,
            port=final_port,
            direction=final_direction,
            learned=True,
            decision_source="MANUAL_NEW",
            justification=build_justification("MANUAL_NEW", flux, nom, anomalies=anomalies),
            anomalie=anomalie_txt,
        )

    add_pending_decision(
        env_name,
        "MANUAL_REQUIRED",
        best_score if best_score else "",
        protocol,
        final_port,
        src_ip,
        final_dst_ip,
        flow_main_sg,
        final_flow_gref,
        final_direction,
        "",
        "",
        # propose une suggestion basée sur la BDD
        suggested_flux=compute_suggestion(df_bdd, protocol, final_port, src_ip, final_dst_ip, flow_main_sg, final_flow_gref)[0],
        suggested_nom=compute_suggestion(df_bdd, protocol, final_port, src_ip, final_dst_ip, flow_main_sg, final_flow_gref)[1],
        ml_modele=get_model_name(best_model_info),
        ml_confiance=str(round(predicted_conf, 4)) if predicted_conf else "",
        seuil_auto=str(round(ml_threshold, 4)) if predicted_conf else "",
        anomalie=anomalie_txt,
    )

    return pack_result(
        "",
        "",
        niveau_securite="Black",
        ml_modele=get_model_name(best_model_info),
        ml_confiance=str(round(predicted_conf, 4)) if predicted_conf else "",
        dst_ip=final_dst_ip,
        flow_gref_sg=final_flow_gref,
        port=final_port,
        direction=final_direction,
        learned=False,
        decision_source="PENDING_MANUAL",
        justification=build_justification("PENDING_MANUAL", "", "", anomalies=anomalies),
        anomalie=anomalie_txt,
    )

# =========================================================
# FR
# =========================================================
def build_fr_file(df_env, df_bdd, env_folder, basicat_code, env_name, idcarto_mapping, best_model_info=None):
    fr_rows = []
    learned_any = False

    for _, row in df_env.iterrows():
        application_name = clean_value(row[COL_APPLICATION])
        src_ip = clean_value(row[COL_IP])
        flow_main_sg = clean_value(row[COL_SGIC])
        app_id_main = clean_value(row[COL_APP_ID])

        protocol = DEFAULT_PROTOCOL
        port_value = ""
        dst_ip = ""
        flow_gref_sg = ""
        direction = DEFAULT_DIRECTION

        resolved = resolve_flow_with_bdd_and_ml(
            df_bdd=df_bdd,
            protocol=protocol,
            port=port_value,
            src_ip=src_ip,
            dst_ip=dst_ip,
            flow_main_sg=flow_main_sg,
            flow_gref_sg=flow_gref_sg,
            direction=direction,
            application=application_name,
            env_name=env_name,
            basicat_code=basicat_code,
            best_model_info=best_model_info,
            ask_user=not API_MODE,
        )

        if resolved["learned"]:
            learned_any = True
            df_bdd = add_bdd_learning_row(
                df_bdd=df_bdd,
                protocol=protocol,
                port=resolved["port"],
                src_ip=src_ip,
                dst_ip=resolved["dst_ip"],
                flow_main_sg=flow_main_sg,
                flow_gref_sg=resolved["flow_gref_sg"],
                direction=resolved["direction"],
                flux=resolved["flux"],
                nom=resolved["nom"],
            )

        app_id_gref = ""
        if resolved["dst_ip"] in idcarto_mapping:
            app_id_gref = idcarto_mapping[resolved["dst_ip"]]

        rule_name = application_name if application_name else f"{basicat_code}_{env_name}"

        fr_rows.append({
            "Name": rule_name,
            "Configured Service": application_name,
            "Status": DEFAULT_STATUS,
            "IP Protocol": protocol,
            "Direction": resolved["direction"],
            "port": resolved["port"],
            "Action": DEFAULT_ACTION,
            "Configured Destination": resolved["dst_ip"],
            "Configured Source": src_ip,
            "flux": resolved["flux"],
            "Nom": resolved["nom"],
            "flowMainSG": flow_main_sg,
            "flowGrefSG": resolved["flow_gref_sg"],
            "applicationIdFlowMainSG": app_id_main,
            "applicationIdFlowGrefSG": app_id_gref,
            "niveau_securite": resolved["niveau_securite"],
            "ml_modele": resolved["ml_modele"],
            "ml_confiance": clean_value(resolved["ml_confiance"]),
            "justification": resolved.get("justification", ""),
            "anomalie": resolved.get("anomalie", ""),
        })

    df_fr = pd.DataFrame(fr_rows, columns=FINAL_COLUMNS)
    fr_file = env_folder / f"{basicat_code}_{env_name}-FR.xlsx"
    # Export FR sans colonnes de label/diagnostic réservées (niveau_securite, ml_*, justification, anomalie)
    fr_export_cols = [c for c in FINAL_COLUMNS if c not in EXPORT_EXCLUDE_LABELS]
    df_fr[fr_export_cols].to_excel(fr_file, index=False)
    autosize_excel_columns(fr_file)
    apply_security_colors(fr_file)

    return df_bdd, learned_any


def run_etape0_and_fr(best_model_info=None):
    excel_path = Path(EXCEL_FILE)
    bdd_path = Path(BDD_FILE)

    if not excel_path.exists():
        print(f"Fichier introuvable : {excel_path}")
        return None, None

    try:
        df = pd.read_excel(excel_path, sheet_name=SHEET_NAME, dtype=str)
    except Exception as e:
        print(f"Erreur lecture fichier Excel principal : {e}")
        return None, None

    df_bdd = load_bdd(bdd_path)
    df_bdd, _ = quality_check_bdd(df_bdd)
    idcarto_mapping = load_idcarto_mapping(excel_path)

    df.columns = [str(col).strip() for col in df.columns]

    required_cols = [COL_BASICAT, COL_PRODUCTION, COL_SGIC, COL_APPLICATION, COL_IP, COL_APP_ID]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print("Colonnes manquantes dans le fichier Excel principal :")
        for col in missing_cols:
            print(f"- {col}")
        return None, None

    for col in required_cols:
        df[col] = df[col].apply(clean_value)

    for col in df_bdd.columns:
        df_bdd[col] = df_bdd[col].apply(clean_value)

    print("=== GENERATION FR TERMINEE ===")
    print("Étape 0 et fichiers FR générés avec succès.")
    print("Le traitement SNIF reste à faire séparément, environnement par environnement.")

    basicat_code = input("Entrez le code BASICAT : ").strip()
    if not basicat_code:
        print("Aucun code BASICAT saisi.")
        return None, None

    df_filtered = df[df[COL_BASICAT].apply(normalize) == normalize(basicat_code)].copy()
    if df_filtered.empty:
        print(f"Aucune ligne trouvée pour BASICAT = {basicat_code}")
        return None, None

    # Déterminer l'environnement en utilisant UTILISATION si elle existe, sinon PRODUCTION
    if COL_UTILISATION in df_filtered.columns:
        df_filtered["ENV"] = df_filtered.apply(
            lambda row: get_environment(row[COL_PRODUCTION], row[COL_UTILISATION]),
            axis=1
        )
        print("✓ Détection d'environnement : UTILISATION + PRODUCTION")
    else:
        df_filtered["ENV"] = df_filtered[COL_PRODUCTION].apply(get_environment)
        print("✓ Détection d'environnement : PRODUCTION uniquement (UTILISATION non trouvée)")
    
    df_filtered = df_filtered[df_filtered["ENV"].notna()].copy()

    if df_filtered.empty:
        print("Aucune ligne exploitable trouvée pour ce BASICAT.")
        print("La colonne PRODUCTION doit contenir 'X'/'x' ou être vide.")
        return None, None

    basicat_folder = ensure_folder(OUTPUT_ROOT / basicat_code)
    generated_envs = []
    learned_any_fr = False

    for env in ["prod", "horsprod"]:
        df_env = df_filtered[df_filtered["ENV"] == env].copy()
        if df_env.empty:
            continue

        generated_envs.append(env)
        env_folder = ensure_folder(basicat_folder / env)

        df_sgic = (
            df_env[[COL_SGIC]]
            .drop_duplicates()
            .sort_values(by=COL_SGIC)
            .reset_index(drop=True)
        )
        sgic_file = env_folder / "sgic.xlsx"
        df_sgic.to_excel(sgic_file, index=False)
        autosize_excel_columns(sgic_file)

        df_app_ip = (
            df_env[[COL_APPLICATION, COL_IP, COL_APP_ID]]
            .drop_duplicates()
            .sort_values(by=[COL_APPLICATION, COL_IP])
            .reset_index(drop=True)
        )
        app_ip_file = env_folder / "applications_ip.xlsx"
        df_app_ip.to_excel(app_ip_file, index=False)
        autosize_excel_columns(app_ip_file)

        df_bdd, learned_fr = build_fr_file(
            df_env=df_env,
            df_bdd=df_bdd,
            env_folder=env_folder,
            basicat_code=basicat_code,
            env_name=env,
            idcarto_mapping=idcarto_mapping,
            best_model_info=best_model_info,
        )
        if learned_fr:
            learned_any_fr = True

    if learned_any_fr:
        save_bdd(df_bdd, bdd_path)
        print("BDD mise à jour après FR.")

    print("\nÉtape 0 + FR terminées avec succès.")
    print(f"Résultat créé dans : {basicat_folder.resolve()}")

    return basicat_code, generated_envs


# =========================================================
# SNIF
# =========================================================
def transform_snif(df_input):
    df = df_input.copy()
    df.columns = [str(c).strip() for c in df.columns]
    
    # Essayer détection automatique des colonnes SNIF
    try:
        from app.services.file_detector import map_columns, SNIF_SIGNATURE
        print("Détection automatique des colonnes SNIF...")
        auto_mapping = map_columns(df, SNIF_SIGNATURE, min_score=50.0)
        if auto_mapping:
            print(f"✓ Colonnes mappées automatiquement : {auto_mapping}")
            df = df.rename(columns=auto_mapping)
    except Exception as e:
        print(f"⚠️  Détection auto échouée, utilisant recherche manuelle: {e}")

    cols_to_drop_exact = [
        "Commentaire",
        "Application",
        "flux",
        "type",
        "Nom",
        "Destination IP Address",
        "Source IP Address",
    ]

    cols_to_drop_dynamic = []
    for c in df.columns:
        c_norm = normalize(c)
        if c_norm.startswith("TRAFFIC RATE (IN BPS)"):
            cols_to_drop_dynamic.append(c)
        elif c_norm.startswith("TOTAL TRAFFIC (IN BYTES)"):
            cols_to_drop_dynamic.append(c)

    for c in cols_to_drop_exact + cols_to_drop_dynamic:
        if c in df.columns:
            df.drop(columns=[c], inplace=True)

    rename_map = {}

    col = find_column_case_insensitive(df.columns, "Destination Security Groups")
    if col:
        rename_map[col] = "flowGrefSG"

    col = find_column_case_insensitive(df.columns, "Source Security Groups")
    if col:
        rename_map[col] = "flowMainSG"

    col = find_column_case_insensitive(df.columns, "Destination IPSets")
    if col:
        rename_map[col] = "Configured Destination"

    col = find_column_case_insensitive(df.columns, "Source IPSets")
    if col:
        rename_map[col] = "Configured Source"

    col = find_column_case_insensitive(df.columns, "port.display")
    if col:
        rename_map[col] = "port"

    col = find_column_case_insensitive(df.columns, ["Protocol", "IP Protocol"])
    if col:
        rename_map[col] = "IP Protocol"

    col = find_column_case_insensitive(df.columns, ["Direction", "direction"])
    if col:
        rename_map[col] = "Direction"

    col = find_column_case_insensitive(df.columns, ["Service", "Configured Service"])
    if col:
        rename_map[col] = "Configured Service"

    col = find_column_case_insensitive(df.columns, ["Name"])
    if col:
        rename_map[col] = "Name"

    df.rename(columns=rename_map, inplace=True)

    for col in FINAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    if "Status" in df.columns:
        df["Status"] = df["Status"].apply(clean_value)
        df.loc[df["Status"] == "", "Status"] = DEFAULT_STATUS
    else:
        df["Status"] = DEFAULT_STATUS

    if "Action" in df.columns:
        df["Action"] = df["Action"].apply(clean_value)
        df.loc[df["Action"] == "", "Action"] = DEFAULT_ACTION
    else:
        df["Action"] = DEFAULT_ACTION

    if "IP Protocol" in df.columns:
        df["IP Protocol"] = df["IP Protocol"].apply(clean_value)
        df.loc[df["IP Protocol"] == "", "IP Protocol"] = DEFAULT_PROTOCOL
    else:
        df["IP Protocol"] = DEFAULT_PROTOCOL

    return df


def enrich_with_bdd(df, df_bdd, env_name="", basicat_code="", idcarto_mapping=None, best_model_info=None):
    learned = False

    stats = {
        "total": 0,
        "exact_match": 0,
        "partial_match": 0,
        "new_rows": 0,
        "validated": 0,
        "corrected": 0,
    }

    detail_rows = []

    if idcarto_mapping is None:
        idcarto_mapping = {}

    for idx, row in df.iterrows():
        stats["total"] += 1

        protocol = clean_value(row.get("IP Protocol", DEFAULT_PROTOCOL)) or DEFAULT_PROTOCOL
        port = clean_value(row.get("port", ""))
        src_ip = clean_value(row.get("Configured Source", ""))
        dst_ip = clean_value(row.get("Configured Destination", ""))
        flow_main_sg = clean_value(row.get("flowMainSG", ""))
        flow_gref_sg = clean_value(row.get("flowGrefSG", ""))
        application = clean_value(row.get("Configured Service", "")) or clean_value(row.get("Name", ""))
        direction = infer_direction(row)

        app_id_main = clean_value(row.get("applicationIdFlowMainSG", ""))
        app_id_gref = clean_value(row.get("applicationIdFlowGrefSG", ""))

        if not app_id_main and src_ip in idcarto_mapping:
            app_id_main = idcarto_mapping[src_ip]

        if not app_id_gref and dst_ip in idcarto_mapping:
            app_id_gref = idcarto_mapping[dst_ip]

        resolved = resolve_flow_with_bdd_and_ml(
            df_bdd=df_bdd,
            protocol=protocol,
            port=port,
            src_ip=src_ip,
            dst_ip=dst_ip,
            flow_main_sg=flow_main_sg,
            flow_gref_sg=flow_gref_sg,
            direction=direction,
            application=application,
            env_name=env_name,
            basicat_code=basicat_code,
            best_model_info=best_model_info,
            ask_user=not API_MODE,
        )

        if resolved["decision_source"] in {"BDD_EXACT", "BDD_PATTERN_AUTO", "HISTORICAL_AUTO"}:
            stats["exact_match"] += 1
            stats["validated"] += 1
        elif resolved["decision_source"] == "BDD_PARTIAL_VALIDATED":
            stats["partial_match"] += 1
            stats["validated"] += 1
        elif resolved["decision_source"] in {"BDD_PARTIAL_CORRECTED", "ML_VALIDATED", "MANUAL_NEW"}:
            stats["corrected"] += 1
            if resolved["decision_source"] == "MANUAL_NEW":
                stats["new_rows"] += 1
            else:
                stats["partial_match"] += 1
        else:
            stats["partial_match"] += 1

        if resolved["learned"]:
            learned = True
            df_bdd = add_bdd_learning_row(
                df_bdd=df_bdd,
                protocol=protocol,
                port=resolved["port"],
                src_ip=src_ip,
                dst_ip=resolved["dst_ip"],
                flow_main_sg=flow_main_sg,
                flow_gref_sg=resolved["flow_gref_sg"],
                direction=resolved["direction"],
                flux=resolved["flux"],
                nom=resolved["nom"],
            )

        if not app_id_main and src_ip in idcarto_mapping:
            app_id_main = idcarto_mapping[src_ip]

        if not app_id_gref and resolved["dst_ip"] in idcarto_mapping:
            app_id_gref = idcarto_mapping[resolved["dst_ip"]]

        if not clean_value(row.get("Name", "")):
            df.at[idx, "Name"] = infer_name(
                application,
                flow_main_sg,
                resolved["flow_gref_sg"],
                resolved["direction"],
                resolved["flux"]
            )

        if not clean_value(row.get("Configured Service", "")):
            df.at[idx, "Configured Service"] = application

        df.at[idx, "Status"] = clean_value(row.get("Status", "")) or DEFAULT_STATUS
        df.at[idx, "IP Protocol"] = protocol
        df.at[idx, "Direction"] = resolved["direction"]
        df.at[idx, "Action"] = clean_value(row.get("Action", "")) or DEFAULT_ACTION
        df.at[idx, "Configured Destination"] = resolved["dst_ip"]
        df.at[idx, "Configured Source"] = src_ip
        df.at[idx, "flux"] = resolved["flux"]
        df.at[idx, "Nom"] = resolved["nom"]
        df.at[idx, "flowMainSG"] = flow_main_sg
        df.at[idx, "flowGrefSG"] = resolved["flow_gref_sg"]
        df.at[idx, "applicationIdFlowMainSG"] = app_id_main
        df.at[idx, "applicationIdFlowGrefSG"] = app_id_gref
        df.at[idx, "niveau_securite"] = resolved["niveau_securite"]
        df.at[idx, "ml_modele"] = resolved["ml_modele"]
        df.at[idx, "ml_confiance"] = clean_value(resolved["ml_confiance"])
        df.at[idx, "justification"] = resolved.get("justification", "")
        df.at[idx, "anomalie"] = resolved.get("anomalie", "")

        detail_rows.append({
            "env": env_name,
            "application": application,
            "src_ip": src_ip,
            "dst_ip": resolved["dst_ip"],
            "flowMainSG": flow_main_sg,
            "flowGrefSG": resolved["flow_gref_sg"],
            "protocol": protocol,
            "port": resolved["port"],
            "flux_final": resolved["flux"],
            "nom_final": resolved["nom"],
            "niveau_securite": resolved["niveau_securite"],
            "decision_source": resolved["decision_source"],
            "ml_modele": resolved["ml_modele"],
            "ml_confiance": clean_value(resolved["ml_confiance"]),
            "justification": resolved.get("justification", ""),
            "anomalie": resolved.get("anomalie", ""),
        })

    print("\n=== EXECUTION ===")
    print(f"Total lignes             : {stats['total']}")
    print(f"Match exact              : {stats['exact_match']}")
    print(f"Match partiel            : {stats['partial_match']}")
    print(f"Nouvelles lignes         : {stats['new_rows']}")
    print(f"Validées                 : {stats['validated']}")
    print(f"Corrigées                : {stats['corrected']}")
    if stats["total"] > 0:
        rate = round((stats["validated"] / stats["total"]) * 100, 2)
        print(f"Taux validation directe  : {rate}%")

    details_df = pd.DataFrame(detail_rows)
    return df, df_bdd, learned, stats, details_df


def process_snif_file(snif_path, bdd_path, basicat_code, env_name, best_model_info=None, benchmark_df=None, conflicts_df=None):
    snif_path = Path(snif_path)
    bdd_path = Path(bdd_path)

    if not snif_path.exists():
        print(f"Fichier SNIF introuvable : {snif_path}")
        return None

    if snif_path.suffix.lower() != ".xlsx":
        print(f"Le fichier doit être un Excel .xlsx : {snif_path}")
        return None

    try:
        df_input = pd.read_excel(snif_path, dtype=str)
    except Exception as e:
        print(f"Erreur lecture fichier SNIF : {e}")
        return None

    df_input.columns = [str(c).strip() for c in df_input.columns]
    df_bdd = load_bdd(bdd_path)
    df_bdd, local_conflicts_df = quality_check_bdd(df_bdd)
    idcarto_mapping = load_idcarto_mapping(Path(EXCEL_FILE))

    df = transform_snif(df_input)

    for idx, row in df.iterrows():
        if not clean_value(row.get("Configured Service", "")):
            df.at[idx, "Configured Service"] = clean_value(row.get("Name", ""))
        if not clean_value(row.get("Name", "")):
            df.at[idx, "Name"] = clean_value(row.get("Configured Service", ""))

    df, df_bdd, learned, stats, details_df = enrich_with_bdd(
        df=df,
        df_bdd=df_bdd,
        env_name=env_name,
        basicat_code=basicat_code,
        idcarto_mapping=idcarto_mapping,
        best_model_info=best_model_info,
    )

    for col in FINAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df_final = df[FINAL_COLUMNS].copy()

    # Export SNIF sans colonnes de label/diagnostic réservées
    snif_export_cols = [c for c in FINAL_COLUMNS if c not in EXPORT_EXCLUDE_LABELS]

    output_path = OUTPUT_ROOT / basicat_code / env_name / f"{safe_filename(basicat_code)}_{env_name}_snif.xlsx"
    ensure_folder(output_path.parent)

    try:
        df_final[snif_export_cols].to_excel(output_path, index=False)
        autosize_excel_columns(output_path)
        apply_security_colors(output_path)
    except Exception as e:
        print(f"Erreur écriture fichier SNIF de sortie : {e}")
        return None

    if learned:
        try:
            save_bdd(df_bdd, bdd_path)
            print(f"BDD mise à jour : {bdd_path.resolve()}")
        except Exception as e:
            print(f"Erreur sauvegarde BDD : {e}")
            return None

    summary = {
        "timestamp": now_timestamp(),
        "basicat": basicat_code,
        "env": env_name,
        "total_lignes": stats["total"],
        "match_exact": stats["exact_match"],
        "match_partiel": stats["partial_match"],
        "nouvelles_lignes": stats["new_rows"],
        "validees": stats["validated"],
        "corrigees": stats["corrected"],
        "modele_ml": get_model_name(best_model_info),
    }

    log_file = write_run_log(
        details_df=details_df,
        summary_dict=summary,
        benchmark_df=benchmark_df,
        conflicts_df=local_conflicts_df if not local_conflicts_df.empty else conflicts_df,
    )

    print(f"Fichier SNIF généré : {output_path.resolve()}")
    print(f"Journal d'exécution   : {log_file.resolve()}")
    return {
        "basicat": basicat_code,
        "envs": [env_name],
        "pending_decisions": list(PENDING_DECISIONS),
        "historical_decisions": list(HISTORICAL_DECISIONS),
        "benchmark": [],
        "conflicts": local_conflicts_df.to_dict(orient="records") if not local_conflicts_df.empty else [],
        "files": [str(output_path)],
    }


# =========================================================
# MATRICE FINALE MAF
# =========================================================
MATRIX_COLUMNS = [
    "nameDeployment", "nameOwner", "cuid", "email",
    "direction IN/OUT", "flowMainSG", "flowGrefSG", "flowGretSG",
    "applicationIdFlowMainSG", "applicationIdFlowGrefSG",
    "protocol TCP/UDP", "port", "justification",
]


def normalize_direction_for_matrix(value):
    value = clean_value(value).lower()
    if value in {"in", "inbound", "entrant", "entrante"}:
        return "IN"
    if value in {"out", "outbound", "sortant", "sortante"}:
        return "OUT"
    return ""


def build_matrix_from_output_df(df_source):
    for col in ["Name", "Direction", "flowMainSG", "flowGrefSG", "applicationIdFlowMainSG", "applicationIdFlowGrefSG", "IP Protocol", "port", "justification"]:
        if col not in df_source.columns:
            df_source[col] = ""
    matrix = pd.DataFrame({
        "nameDeployment": df_source["Name"].apply(clean_value),
        "nameOwner": "",
        "cuid": "",
        "email": "",
        "direction IN/OUT": df_source["Direction"].apply(normalize_direction_for_matrix),
        "flowMainSG": df_source["flowMainSG"].apply(clean_value),
        "flowGrefSG": df_source["flowGrefSG"].apply(clean_value),
        "flowGretSG": df_source["flowGrefSG"].apply(clean_value),
        "applicationIdFlowMainSG": df_source["applicationIdFlowMainSG"].apply(clean_value),
        "applicationIdFlowGrefSG": df_source["applicationIdFlowGrefSG"].apply(clean_value),
        "protocol TCP/UDP": df_source["IP Protocol"].apply(lambda v: normalize(v) if normalize(v) in {"TCP", "UDP"} else ""),
        "port": df_source["port"].apply(clean_value),
        "justification": df_source["justification"].apply(clean_value),
    })
    return matrix.drop_duplicates().reset_index(drop=True)[MATRIX_COLUMNS]


def generate_final_maf_matrix(basicat_code, env_name):
    env_folder = OUTPUT_ROOT / basicat_code / env_name
    if not env_folder.exists():
        return None
    source_files = []
    fr_file = env_folder / f"{basicat_code}_{env_name}-FR.xlsx"
    snif_file = env_folder / f"{safe_filename(basicat_code)}_{env_name}_snif.xlsx"
    if fr_file.exists():
        source_files.append(fr_file)
    if snif_file.exists():
        source_files.append(snif_file)
    frames = []
    for f in source_files:
        try:
            frames.append(pd.read_excel(f).fillna(""))
        except Exception as e:
            print(f"Impossible de lire {f}: {e}")
    if not frames:
        return None
    df_all = pd.concat(frames, ignore_index=True)
    matrix = build_matrix_from_output_df(df_all)
    out_file = env_folder / f"{safe_filename(basicat_code)}_{env_name}_MAF.xlsx"
    matrix.to_excel(out_file, index=False)
    autosize_excel_columns(out_file)
    print(f"Matrice MAF générée : {out_file.resolve()}")
    return out_file


def _prompt_snif_from_applications_ip(basicat_code: str, env: str) -> str:
    """
    Affiche le contenu de applications_ip.xlsx pour `basicat_code`/`env`,
    génère la chaîne de recherche (« Flow in last 31 days where VM = ... »),
    affiche le message à copier et demande le chemin du fichier SNIF.
    Retourne la chaîne fournie par l'utilisateur (vide si ignoré).
    """
    base = OUTPUT_ROOT / basicat_code
    app_ip = base / env / "applications_ip.xlsx"

    # Fallback: try relative output_basicat (compat workspace layouts)
    if not app_ip.exists():
        alt = Path("output_basicat") / basicat_code / env / "applications_ip.xlsx"
        if alt.exists():
            app_ip = alt

    if not app_ip.exists():
        # no applications_ip available; fallback to simple prompt
        return input(f"\nChemin fichier SNIF {env.upper()} : ").strip().replace('"', "")

    try:
        df = pd.read_excel(app_ip, dtype=str).fillna("")
    except Exception as e:
        print(f"Impossible de lire {app_ip}: {e}")
        return input(f"\nChemin fichier SNIF {env.upper()} : ").strip().replace('"', "")

    # Détecter colonnes name / ip / idcarto de façon tolérante
    name_col = None
    ip_col = None
    id_col = None
    for c in df.columns:
        lc = str(c).strip().lower()
        if not name_col and lc in {"name", "configured service", "application"}:
            name_col = c
        if not ip_col and lc in {"ip", "ip address", "ipaddr"}:
            ip_col = c
        if not id_col and lc in {"idcarto", "applicationid", "id"}:
            id_col = c

    if name_col is None and len(df.columns) > 0:
        name_col = df.columns[0]

    display_cols = [c for c in (name_col, ip_col, id_col) if c]
    print("\n" + "="*60)
    print(f"Environnement: {env} — fichier: {app_ip}")
    print("="*60)
    if display_cols:
        try:
            print(df[display_cols].to_string(index=False))
        except Exception:
            print(df.to_string(index=False))
    else:
        print(df.to_string(index=False))

    # Construire la chaîne à copier
    try:
        names = df[name_col].dropna().astype(str).tolist()
    except Exception:
        names = []

    names = [n.replace('"', '') for n in names if n.strip()]
    if names:
        query = 'Flow in last 31 days where VM = ' + '"' + '" OR VM = "'.join(names) + '"'
    else:
        query = ""

    print("\n--- MESSAGE A COPIER (coller dans l'outil de recherche) ---")
    if query:
        print(query)
    else:
        print("(aucune VM trouvée dans applications_ip)")
    print("--- FIN MESSAGE ---\n")

    # Option de passer rapidement cette étape
    choice = input(f"Tapez 'p' puis Entrée pour PASSER {env.upper()}, ou appuyez sur Entrée pour continuer : ").strip().lower()
    if choice == 'p':
        print(f"Étape SNIF {env.upper()} passée par l'utilisateur.")
        return ""

    input(f"Copiez le message ci-dessus. Appuyez sur Entrée quand vous êtes prêt à déposer le fichier SNIF pour {env.upper()}.")
    snif_path = input(f"Chemin du fichier SNIF {env.upper()} (laisser vide pour sauter) : ").strip()
    return snif_path.replace('"', "")

# =========================================================
# MAIN
# =========================================================
def generate_applications_ip_preview(basicat_code: str, env: str) -> dict:
    """Return a preview dict for applications_ip.xlsx for UI consumption.
    Returns: { path: str, columns: [...], head: [row dicts], query: str }
    """
    base = OUTPUT_ROOT / basicat_code
    app_ip = base / env / "applications_ip.xlsx"

    # Fallback: try relative output_basicat
    if not app_ip.exists():
        alt = Path("output_basicat") / basicat_code / env / "applications_ip.xlsx"
        if alt.exists():
            app_ip = alt

    if not app_ip.exists():
        return {"path": "", "columns": [], "head": [], "query": ""}

    try:
        df = pd.read_excel(app_ip, dtype=str).fillna("")
    except Exception:
        return {"path": str(app_ip), "columns": [], "head": [], "query": ""}

    # detect columns
    name_col = None
    ip_col = None
    id_col = None
    for c in df.columns:
        lc = str(c).strip().lower()
        if not name_col and lc in {"name", "configured service", "application", "vm", "vm_name"}:
            name_col = c
        if not ip_col and lc in {"ip", "ip address", "ipaddr"}:
            ip_col = c
        if not id_col and lc in {"idcarto", "applicationid", "id"}:
            id_col = c

    if name_col is None and len(df.columns) > 0:
        name_col = df.columns[0]

    display_cols = [c for c in (name_col, ip_col, id_col) if c]

    # build query
    try:
        names = df[name_col].dropna().astype(str).tolist() if name_col else []
    except Exception:
        names = []
    names = [n.replace('"', '') for n in names if n.strip()]
    query = 'Flow in last 31 days where VM = ' + '"' + '" OR VM = "'.join(names) + '"' if names else ""

    head_records = []
    try:
        head_records = df[display_cols].head(15).to_dict(orient="records") if display_cols else df.head(15).to_dict(orient="records")
    except Exception:
        head_records = df.head(15).to_dict(orient="records")

    return {
        "path": str(app_ip),
        "columns": [str(c) for c in df.columns],
        "head": head_records,
        "query": query,
    }
def main():
    df_bdd_for_benchmark = load_bdd(Path(BDD_FILE))
    df_bdd_for_benchmark, conflicts_df = quality_check_bdd(df_bdd_for_benchmark)

    if not conflicts_df.empty:
        print("ATTENTION : des conflits ont été détectés dans la BDD.")
        print(f"Nombre de signatures conflictuelles : {len(conflicts_df)}")
        print("Un rapport sera exporté dans le journal d'exécution.\n")

    best_model_info = None
    benchmark_df = pd.DataFrame()

    basicat_code, generated_envs = run_etape0_and_fr(best_model_info=best_model_info)
    if not basicat_code:
        return

    bdd_path = Path(BDD_FILE)

    print("\n=== ETAPE SNIF ===")
    print("Tu peux maintenant drag & drop le fichier Excel SNIF de prod et/ou horsprod.")
    print("Si tu n'en as pas un, appuie juste sur Entrée pour le sauter.")

    if "prod" in generated_envs:
        snif_prod = _prompt_snif_from_applications_ip(basicat_code, "prod")
        if snif_prod:
            process_snif_file(
                snif_path=snif_prod,
                bdd_path=bdd_path,
                basicat_code=basicat_code,
                env_name="prod",
                best_model_info=best_model_info,
                benchmark_df=benchmark_df,
                conflicts_df=conflicts_df,
            )

    if "horsprod" in generated_envs:
        snif_horsprod = _prompt_snif_from_applications_ip(basicat_code, "horsprod")
        if snif_horsprod:
            process_snif_file(
                snif_path=snif_horsprod,
                bdd_path=bdd_path,
                basicat_code=basicat_code,
                env_name="horsprod",
                best_model_info=best_model_info,
                benchmark_df=benchmark_df,
                conflicts_df=conflicts_df,
            )

    for env in generated_envs:
        generate_final_maf_matrix(basicat_code, env)

    print("\nTraitement complet terminé.")


def run_maf(basicat_code, snif_prod=None, snif_horsprod=None, finalize_maf=True):
    """
    Version API-safe du moteur MAF.
    - Ne bloque pas FastAPI avec input().
    - Utilise les chemins backend depuis app.core.config.
    - Les cas ambigus sont ajoutés dans PENDING_DECISIONS et retournés au front.
    """
    global API_MODE, PENDING_DECISIONS, HISTORICAL_DECISIONS, input

    API_MODE = True
    PENDING_DECISIONS = []
    HISTORICAL_DECISIONS = []

    df_bdd_for_benchmark = load_bdd(Path(BDD_FILE))
    df_bdd_for_benchmark, conflicts_df = quality_check_bdd(df_bdd_for_benchmark)
    best_model_info = None
    benchmark_df = pd.DataFrame()

    original_input = input

    def fake_basicat_input(prompt=""):
        if "BASICAT" in str(prompt).upper():
            return str(basicat_code).strip().upper()
        return ""

    input = fake_basicat_input

    try:
        resolved_basicat, generated_envs = run_etape0_and_fr(best_model_info=best_model_info)
    finally:
        input = original_input

    if not resolved_basicat:
        API_MODE = False
        return {
            "basicat": str(basicat_code).strip().upper(),
            "envs": [],
            "pending_decisions": PENDING_DECISIONS,
            "historical_decisions": HISTORICAL_DECISIONS,
            "files": [],
            "error": "Aucun BASICAT traité",
        }

    bdd_path = Path(BDD_FILE)

    if snif_prod and "prod" in generated_envs:
        process_snif_file(
            snif_path=snif_prod,
            bdd_path=bdd_path,
            basicat_code=resolved_basicat,
            env_name="prod",
            best_model_info=best_model_info,
            benchmark_df=benchmark_df,
            conflicts_df=conflicts_df,
        )

    if snif_horsprod and "horsprod" in generated_envs:
        process_snif_file(
            snif_path=snif_horsprod,
            bdd_path=bdd_path,
            basicat_code=resolved_basicat,
            env_name="horsprod",
            best_model_info=best_model_info,
            benchmark_df=benchmark_df,
            conflicts_df=conflicts_df,
        )

    generated_files = []
    if finalize_maf:
        for env in generated_envs:
            generate_final_maf_matrix(resolved_basicat, env)
            env_folder = OUTPUT_ROOT / resolved_basicat / env
            if env_folder.exists():
                for f in env_folder.rglob("*.xlsx"):
                    generated_files.append(str(f))
    else:
        for env in generated_envs:
            env_folder = OUTPUT_ROOT / resolved_basicat / env
            if env_folder.exists():
                for f in env_folder.rglob("*.xlsx"):
                    generated_files.append(str(f))

    API_MODE = False

    return {
        "basicat": resolved_basicat,
        "envs": generated_envs,
        "pending_decisions": PENDING_DECISIONS,
        "historical_decisions": HISTORICAL_DECISIONS,
        "files": generated_files,
        "benchmark": benchmark_df.to_dict(orient="records") if benchmark_df is not None and not benchmark_df.empty else [],
        "conflicts": conflicts_df.to_dict(orient="records") if conflicts_df is not None and not conflicts_df.empty else [],
    }
